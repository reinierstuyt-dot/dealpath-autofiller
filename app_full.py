from flask import Flask, request, jsonify
import pandas as pd, openpyxl, requests, io, re, datetime, os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Kleuren voor markeringen
GREY = PatternFill(start_color="00D9D9D9", end_color="00D9D9D9", fill_type="solid")
YELLOW = PatternFill(start_color="00FFF2CC", end_color="00FFF2CC", fill_type="solid")

app = Flask(__name__)

# Regels per sheet
RULES = {
  "Leasing Comps ": {
    "self": ["Property Name","Property Type","Comp Type"],
    "required": ["Year","Quarter","Signed Date","Country Code","Country","Main City","Submarket",
                 "Tenant","In-Place Rent (€/sqm)","Term (years)","GLA (sqm)","Built Year",
                 "Address","date added to database"],
    "nice": ["Landlord","Rent Free (months)","% Office","Height (m)","Doors","Cross Dock? (Yes / No)","Location Link","Commentary"],
    "manual": ["Source"]
  },
  "Leasing Supply": {
    "self": ["Property Name","Property Type","Comp Type"],
    "required": ["Year","Quarter","Signed Date","Country Code","Country","Main City","Submarket",
                 "Asking Rent (€/sqm)","GLA (sqm)","Built Year","Address","date added to database"],
    "nice": ["Landlord","Term (years)","Incentives (months)","% Office","Height (m)","Doors","Cross Dock? (Yes / No)","Location Link","Commentary"],
    "manual": ["Source"]
  },
  "Investment Comps": {
    "self": ["Property Name","Property Type","Comp Type"],
    "required": ["Year","Quarter","Signed Date","Country Code","Country","Main City","Submarket",
                 "Net PP","Net PP /psm","NIY (IP NOI / AIC)","Purchaser","Vendor","Number of Assets","GLA (sqm)",
                 "Address","date added to database"],
    "nice": ["Occupancy (%)","Sale And Leaseback","MTM (MTM NOI / AIC + Capex)","% office","Height (m)","Doors","Year Built","Cross Dock? (Yes / No)","Location Link","Commentary"],
    "manual": ["Source"]
  },
  "Investment Supply": {
    "self": ["Property Name","Property Type","Comp Type"],
    "required": ["Year","Quarter","Signed Date","Country Code","Country","Main City","Submarket",
                 "Vendor","Number of Assets","GLA (sqm)","Address","date added to database"],
    "nice": ["Purchaser (if known)","Asking Price (Net PP)","Asking Price /psm","Indicative NIY","% office","Height (m)","Doors","Year Built","Cross Dock? (Yes / No)","Location Link","Commentary","Occupancy (%)"],
    "manual": ["Source"]
  }
}

# Helpers ---------------------------------------------------

def download(url: str) -> bytes:
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    return r.content

def parse_date(text: str) -> str:
    for pat in [r'(\d{1,2}\s+\w+\s+\d{4})', r'(\w+\s+\d{4})', r'(\d{4}-\d{2}-\d{2})', r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})']:
        m = re.search(pat, text, flags=re.I)
        if m:
            try:
                dt = pd.to_datetime(m.group(1), dayfirst=True).date()
                return dt.isoformat()
            except:
                continue
    return ""

def derive_yq(iso_date: str):
    if not iso_date:
        return "", ""
    d = pd.to_datetime(iso_date, errors="coerce")
    if pd.isna(d):
        return "", ""
    return int(d.year), int((d.month-1)//3 + 1)

# -----------------------------------------------------------

def append_row_preserve(wb_path: str, sheet_name: str, record: dict, defaults: dict):
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[2]]
    header_index = {h: i+1 for i,h in enumerate(headers)}

    # Build new row
    filled = {}
    for h in headers:
        val = record.get(h, "")
        if not val and defaults:
            val = defaults.get(h, "")
        filled[h] = val

    # Derive year/quarter
    if "Signed Date" in filled and filled["Signed Date"]:
        try:
            d = pd.to_datetime(filled["Signed Date"], errors="coerce")
            if not pd.isna(d):
                filled["Year"] = int(d.year)
                filled["Quarter"] = int((d.month-1)//3 + 1)
                filled["Signed Date"] = d.date().isoformat()
        except:
            pass

    if "date added to database" in headers and not filled.get("date added to database"):
        filled["date added to database"] = datetime.date.today().isoformat()

    r = RULES[sheet_name]
    required = set(r.get("required", []))
    manual = set(r.get("manual", []))

    last_row = ws.max_row + 1
    for h, col_idx in header_index.items():
        value = filled.get(h, "")
        ws.cell(row=last_row, column=col_idx, value=value)

        if h in required and (value in ["", None]):
            ws.cell(row=last_row, column=col_idx).fill = GREY
        if h in manual:
            ws.cell(row=last_row, column=col_idx).fill = YELLOW

    tmp_out = wb_path.replace(".xlsx", f"_UPDATED_{sheet_name.replace(' ','_')}.xlsx")
    wb.save(tmp_out)
    return tmp_out

# -----------------------------------------------------------

@app.route("/ingest_and_append", methods=["POST"])
def ingest_and_append():
    data = request.get_json(force=True)
    template_url = data["template_url"]
    sheet_type = data["sheet_type"]
    sheet_status = data["sheet_status"]
    text = data.get("text","")
    defaults = data.get("defaults", {})

    # Download template
    content = download(template_url)
    tmp_in = "template_in.xlsx"
    with open(tmp_in, "wb") as f:
        f.write(content)

    # Choose sheet
    if sheet_type == "leasing" and sheet_status == "comps":
        sheet_name = "Leasing Comps "
    elif sheet_type == "leasing" and sheet_status == "supply":
        sheet_name = "Leasing Supply"
    elif sheet_type == "investment" and sheet_status == "comps":
        sheet_name = "Investment Comps"
    elif sheet_type == "investment" and sheet_status == "supply":
        sheet_name = "Investment Supply"
    else:
        return jsonify({"error": "Invalid sheet_type/status"}), 400

    # Parse record (vereenvoudigd)
    rec = {}
    date_iso = parse_date(text)
    if date_iso: rec["Signed Date"] = date_iso
    y,q = derive_yq(date_iso)
    rec["Year"], rec["Quarter"] = y, q

    # Append
    out_path = append_row_preserve(tmp_in, sheet_name, rec, defaults)

    return jsonify({
        "sheet_name": sheet_name,
        "rows_appended": 1,
        "updated_file_url": out_path
    })

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8080"))
    app.run(host="0.0.0.0", port=port, debug=False)
